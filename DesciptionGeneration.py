import xlrd
import pandas as pd
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
nltk.set_proxy('https://genproxy:8080')
nltk.download('punkt')
import bs4 as bs
import urllib.request
from nltk.tokenize import sent_tokenize, word_tokenize
nltk.download('stopwords')
import heapq
from openpyxl import load_workbook


'''Read the data from excel'''
FileName='Generic4.xlsx'
df = pd.read_excel(FileName,sheet_name='Calendar Business Coverage')
ReadList=df.columns.values.tolist()
#print(ReadList)
#for i in range(0,9) :
#print(df['Last updated on: '][6:])
#print(df['Unnamed: 2'])
#print(df['08 Oct 2018 16:26'][5])

DataColumn=ReadList[0]
ParameterNameCol=df[ReadList[1]][6:].values.tolist()
#print(ParameterNameCol)

ReadListLength=len(ReadList)
#print(ReadListLength)

list1 = df[DataColumn][6:].values.tolist()
#print (list1)


 '''Clean the list i.e removal of NaN'''
cleanlist1 = [x for x in list1 if str(x)!='nan']
#print (cleanlist1)
str1 = ""
str1 = ' '.join(cleanlist1)
#print (str1)


'''Removing the unwanted numbers from the text'''
import re
article_text = re.sub(r'\[[0-9]*\]', ' ', str1)  
article_text = re.sub(r'\s+', ' ', str1) 
article_text = re.sub(r'[0-9]+\.', " ", str1)
#article_text = re.sub("^\d+\s|\s\d+\s|\s\d+$", " ", str1)
#article_text = " ".join([x for x in str1.split(" ") if not x.isdigit()])
#result = ''.join([i for i in str1 if not i.isdigit()])
#print (article_text)
#print (result)

'''Creating Formatted text for analysis purpose '''
formatted_article_text = re.sub('[^a-zA-Z]', ' ', article_text )  
formatted_article_text = re.sub(r'\s+', ' ', formatted_article_text)  
#print (formatted_article_text)


'''Tokenize the sentences'''
sentence_list = nltk.sent_tokenize(article_text) 
#print (sentence_list)

'''Compute word frequencies'''
stopwords = nltk.corpus.stopwords.words('english')
word_frequencies = {}  
for word in nltk.word_tokenize(formatted_article_text):  
    if word not in stopwords:
        print(word)
        if word not in word_frequencies.keys():
            word_frequencies[word] = 1
        else:
            word_frequencies[word] += 1


#print (word_frequencies)


'''Calculate sentence scores'''
sentence_scores = {}  
for sent in sentence_list:  
    for word in nltk.word_tokenize(sent.lower()):
        if word in word_frequencies.keys():
            #if len(sent.split(' ')) < 30:
            if sent not in sentence_scores.keys():
                sentence_scores[sent] = word_frequencies[word]
            else:
                sentence_scores[sent] += word_frequencies[word]

#print (sentence_scores)



'''Extract the parametrs from the excel'''

parameterName = df[ReadList[1]][9:].values.tolist()
#print (parameterName)
   
parameterValue1=df[ReadList[2]][9:].values.tolist()
#print(parameterValue1)
   
parameterNameList = [x for x in parameterName if x != '[Step Data]']
print(parameterNameList)
    
parameterValue1 = [x for x in parameterValue1 if x != 'Step Data :']
#print(parameterValue1)
    
parameterValue2=df[ReadList[3]][9:].values.tolist()
#rint(parameterValue2)
    
parameterValue2= [x for x in parameterValue2 if x != 'Step Data :']
#print(parameterValue2)


'''Generate the Summary'''
summary_sentences = heapq.nlargest(3, sentence_scores, key=sentence_scores.get)
summary = ' '.join(summary_sentences) 
wb=load_workbook(FileName)
#print(wb.sheetnames)
sheet=wb['Calendar Business Coverage']

scenario1=parameterName[0]+": "+parameterValue1[0]
scenario2=parameterName[0]+": "+parameterValue2[0]
'''
print("Scenario1")
print(summary)
print("Parameters : "+scenario1)
print("************************************************************************************************")
print("Scenario2")
print(summary)
print("Parameters : "+scenario2)
'''
'''

print("*****************************************************************************************************")
i='F'
for i in range(2,len(ReadList)):
    parameterValue1=df[ReadList[i]][9:].values.tolist()
    parameterValue1 = [x for x in parameterValue1 if x != 'Step Data :']
    print(summary)
    for i in range(0,len(parameterNameList)):
        scn1=parameterNameList[i]+": "+parameterValue1[i]
        print(scn1)
    
        
    print("**********************************************************************************************")
    
'''
x='F'
index=0
summary=summary+"\n"
for i in range(2,len(ReadList)):
    parameterValue1=df[ReadList[i]][9:].values.tolist()
    parameterValue1 = [x for x in parameterValue1 if x != 'Step Data :']
    print(summary)
    li=[]
    for i in range(0,len(parameterNameList)):
        
        scn1=parameterNameList[i]+": "+parameterValue1[i]
        print(scn1)
        li.insert(index,scn1)
        #print(li)
        index=index+1
    #print(li) 
    str1 = ' , '.join(li)
    cellNumber=""
    cellNumber=x+"6"
    print(cellNumber)
    sheet[cellNumber]=summary+str1
    number=ord(x)
    number=number+1
    x=chr(number)
    wb.save(FileName)
    
    
        
    print("**********************************************************************************************")
    
    
        
    print("**********************************************************************************************")
  
